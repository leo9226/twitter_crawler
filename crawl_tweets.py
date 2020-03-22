import tweepy
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import datetime
import time
from typing import NewType
dt = NewType('dt', datetime)
from src.static import setup_custom_logger
from src.static import move_sheet

# TODO: Filter out special characters (smileys..), hashtags, links
# TODO: If tweets have been crawled already for the given period & account, do not execute


class CrawlTweets:
    def __init__(self, consumer_key: str, consumer_secret: str, key: str, secret: str, path:str):
        self.consumer_key = consumer_key
        self.consumer_secret = consumer_secret
        self.key = key
        self.secret = secret
        self.path = path
        self.logger = setup_custom_logger('Bahn')
        self.api = self.establish_connection()

    def establish_connection(self):
        # Authenticate to Twitter
        auth = tweepy.OAuthHandler(consumer_key=self.consumer_key, consumer_secret=self.consumer_secret)
        auth.set_access_token(key=self.key, secret=self.secret)

        # Create API object
        api = tweepy.API(auth)
        self.logger.info(f'{type(self).__name__} instantiated. API connection successfully established.')
        return api

    def save_tweets(self, tweets: pd.DataFrame, start_date: dt, end_date: dt):
        self.logger.info(f'Saving tweets from {tweets.name} to an Excel file.')
        random = '_' + str(time.time())[-3:]
        file_name = fr'{tweets.name[1:]}.xlsx'
        sheet_name = f'{str(start_date)[:10]}_to_{str(end_date)[:10]}_{random}'
        full_path = os.path.join(self.path, file_name)
        # file_name = fr'{tweets.name[1:]}' + '_' + str(start_date)[:10] + '_to_' + str(end_date)[:10] + random + '.csv'
        if not os.path.exists(self.path):
            os.mkdir(self.path)

        if not os.path.isfile(f'{full_path}'):
            tweets.to_excel(full_path, index=False, sheet_name=sheet_name)
            writer = pd.ExcelWriter(full_path, engine='openpyxl')
            book = load_workbook(full_path)
            writer.book = book
            tweets.to_excel(writer, index=False, sheet_name='main_sheet')
            move_sheet(workbook=book)
            writer.save()
            writer.close()
        else:
            original = pd.read_excel(full_path, 'main_sheet')
            reduced = pd.concat([original, tweets], sort=False)\
                .sort_values(by='date', ascending=False).drop_duplicates(subset=['date', 'tweet'])
            writer = pd.ExcelWriter(full_path, engine='openpyxl')
            book = load_workbook(full_path)
            writer.book = book
            remove = book.get_sheet_by_name('main_sheet')
            book.remove_sheet(remove)
            tweets.to_excel(writer, index=False, sheet_name=sheet_name)
            reduced.to_excel(writer, index=False, sheet_name='main_sheet')
            move_sheet(workbook=book)
            writer.save()
            writer.close()

    def extract_info(self, raw_tweets: list) -> pd.DataFrame:
        self.logger.info(f'Extracting information from raw tweets.')
        created_at, text, lang, hashtags, retweeted, retweet_count, user_mentions = [], [], [], [], [], [], []
        for tweet in raw_tweets:
            created_at.append(tweet.created_at)
            text.append(tweet.text)
            lang.append(tweet.lang)
            hashtags.append([bla['text'] if isinstance(bla['text'], str) else '' for bla in tweet.entities['hashtags']])
            # geo.append(tweet.geo)
            # place.append(tweet.place)
            retweeted.append(tweet.text.startswith('RT '))
            retweet_count.append(tweet.retweet_count)
            user_mentions.append([bla['screen_name'] if isinstance(bla['screen_name'], str) else '' for bla in tweet.entities['user_mentions']])
        df = pd.DataFrame({
            'date': created_at, 'tweet': text, 'hashtags': hashtags, 'language': lang, 'retweeted': retweeted,
            'retweeted_count': retweet_count, 'user_mentions': user_mentions
        })
        df['date'] = pd.to_datetime(df['date'])
        df['tweet'] = df['tweet'].astype(str)
        return df

    def get_tweets(self, user: str, start_date: dt, end_date: dt) -> pd.DataFrame:
        self.logger.info(f'Crawling tweets for {user}.')

        raw_tweets = []
        db = self.api.get_user(user)
        tmpTweets = db.timeline()
        for tweet in tmpTweets:
            if end_date > tweet.created_at > start_date:
                raw_tweets.append(tweet)

        while tmpTweets[-1].created_at > start_date:
            tmpTweets = db.timeline(max_id=tmpTweets[-1].id)
            for tweet in tmpTweets:
                if end_date > tweet.created_at > start_date:
                    raw_tweets.append(tweet)

        extracted = self.extract_info(raw_tweets=raw_tweets)
        extracted.name = user
        self.save_tweets(tweets=extracted, start_date=start_date, end_date=end_date)
        self.logger.info(f'Tweets until {end_date} have been crawled.')

        return extracted

    def crawl_wrapper(self, names: list, start_dates: list, end_dates: list) -> dict:
        if len(str(start_dates[0])) == 4:
            start_dates, end_dates = [start_dates], [end_dates]
        assert len(names) == len(start_dates) == len(end_dates), 'Length of inputs must be the same!'

        crawled_tweets = {}
        for name, start, end in zip(names, start_dates, end_dates):
            try:
                if start >= end:
                    self.logger.info('Start date bigger than end date. Fix dates. Nothing will be crawled.')
                start_date = datetime.datetime(start[0], start[1], start[2], 0, 0)
                end_date = datetime.datetime(end[0], end[1], end[2], 23, 59)
                tweets = self.get_tweets(user=name, start_date=start_date, end_date=end_date)
                crawled_tweets[name] = tweets
            except Exception as e:
                if e.__class__.__qualname__ == 'TweepError':
                    self.logger.info(f'Error code: {e.args[0][0]["code"]}, Error message: {e.args[0][0]["message"]}')
                else:
                    self.logger.info(e)

        return crawled_tweets


ct = CrawlTweets(
    consumer_key="consuemer_key",
    consumer_secret="customer_secret",
    key="key",
    secret="secret",
    path='custom_path'
)

ct.crawl_wrapper(
    names=['@realDonaldTrump', '@Regio_NRW', '@streckenagentOB', '@elonmusk'],
    start_dates=[[2020, 3, 14], [2020, 3, 14], [2020, 3, 1], [2020, 3, 1]],
    end_dates=[[2020, 3, 21], [2020, 3, 21], [2020, 3, 21], [2020, 3, 21]]
)
